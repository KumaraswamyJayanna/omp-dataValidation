"""To generate a category level and file level summary reports"""
import os.path
import os
import shutil
import re
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime

import pandas as pd

from config import ACCURACYTHRESHOLD, GTPATH, OUTPUTPATH, REPORTPATH


report = 'Reports\highlighted_report_pipelineValidationData_result_temp20250219_221316.xlsx'


class File_Report:

    def __init__(self, reportpath, pipelinefile) -> None:
        self.reportpath = reportpath
        self.pipelinefile = pipelinefile
        self.df_report = pd.read_excel(self.reportpath)
        self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.summaryreport = f'summary_{self.timestamp}.xlsx'
        self.counts = {}


    def get_file_names(self):
        filenames = self.df_report['File_Name'].unique()
        return filenames


    def count_column_highlights_ofreport(self, ):
        wb = openpyxl.load_workbook(self.reportpath)
        sheet =wb["Pipeline"]
        highlighted_counts={}

        for col in sheet.columns:
            col_name = col[0].value  # Get column letter (e.g., 'A', 'B', ...)
            highlighted_count = 0

            # Check each cell in the column
            for cell in col:
                if cell.fill and cell.fill.start_color.index != '00000000':  # Check if fill is not default (No color)
                    highlighted_count += 1

            # Store the count of highlighted cells for this column
            highlighted_counts[col_name] = highlighted_count

        return highlighted_counts

    def count_column_highlights_byfile(self, data_file):
        wb = openpyxl.load_workbook(data_file)
        sheet =wb["Sheet"]
        highlighted_counts={}

        for col in sheet.columns:
            col_name = col[0].value  # Get column letter (e.g., 'A', 'B', ...)
            highlighted_count = 0

            # Check each cell in the column
            for cell in col:
                if cell.fill and cell.fill.start_color.index != '00000000':  # Check if fill is not default (No color)
                    highlighted_count += 1

            # Store the count of highlighted cells for this column
            highlighted_counts[col_name] = highlighted_count

        return highlighted_counts

    def get_columns(self):
        return self.df_report.columns

    def filter_by_category(self, column_name, filter_value, output_file_path):

        wb = openpyxl.load_workbook(self.reportpath)
        sheet =wb["Pipeline"]
        filtered_data =[]

        # Get the column index based on the header
        header_row = sheet[1]  # Assuming the first row contains headers
        column_index = None
        for idx, cell in enumerate(header_row, 1):
            if cell.value == column_name:
                column_index = idx
                break

        if column_index is None:
            print(f"Column '{column_name}' not found.")
            return

        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active

        # Copy header row to the new sheet
        for col in range(1, len(header_row) + 1):
            new_sheet.cell(row=1, column=col, value=header_row[col - 1].value)

        filtered_row_count = 0
        # Loop through the rows in the column (skip the first row which is the header)
        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=len(header_row)):
            cell = row[column_index-1]  # Get the cell in the column

            # Check if the cell matches the filter value and if it's highlighted
            if cell.value == filter_value:
                filtered_row_count += 1
                for col_idx, col_cell in enumerate(row, 1):
                    new_cell = new_sheet.cell(row=filtered_row_count + 1, column=col_idx, value=col_cell.value)
                    # Preserve highlighting (if any)
                    if col_cell.fill and col_cell.fill.start_color.index != '00000000':
                        new_cell.fill = PatternFill(start_color=col_cell.fill.start_color.index,
                                                end_color=col_cell.fill.end_color.index,
                                                fill_type=col_cell.fill.fill_type
                                                )
        # Save the new workbook with filtered data
        if filtered_row_count > 0:
            new_wb.save(output_file_path)
            print(f"Filtered data saved to {output_file_path}")
        else:
            print(f"No rows found with '{column_name}' equal to '{filter_value}'.")
        temp_file_path = os.path.abspath(output_file_path)
        return filtered_data, temp_file_path

    def get_count_totaldata(self):
        df_pipeline = pd.read_excel(self.pipelinefile)
        return len(df_pipeline)

    def count_incorrects(self, file_report:pd.DataFrame):
        # Count the sum of values in each column and return as a dictionary
        column_sum = file_report.apply(lambda x: x.sum()).to_dict()
        return column_sum

    def find_by_files(self):
        filenames= self.get_file_names()
        if not os.path.exists("temp"):
            os.makedirs("temp")

        for file in filenames:
            outpath = f'temp/{file}_temp.xlsx'
            _, filtereddatapath = report.filter_by_category("File_Name", file, outpath)
            res = report.count_column_highlights_byfile(filtereddatapath)
            self.counts[file]=res
        df_result=pd.DataFrame(self.counts).T
        return df_result


if __name__=="__main__":
# def analyze_summary_report():

    # Generate a file level report
    report = File_Report(report, OUTPUTPATH)
    df_results = report.find_by_files()


    # Generate category level report
    totalrows = report.get_count_totaldata()
    out = report.count_incorrects(df_results)
    accuracy ={}

    for field, count in out.items():
        accuracy[field]=round(100-((count/totalrows)*100),2)

    # Check number of files affected
    files_affected = df_results != 0
    files_affected_counts = files_affected.sum()

    df_category = pd.DataFrame(
        columns=["Iteration Number", "Issue Type", "Issue Level", "Overall Accuracy Percentage",
                "Number of Files affected"])
    measures = list(df_results.columns)
    for measure in measures:
        category={
            "Iteration Number": "0.0",
            "Issue Type" : measure,
            "Issue Level" : "Field",
            "Overall Accuracy Percentage": float(accuracy[measure]),
            "Number of Files affected": float(files_affected_counts[measure])
        }

        df_category.loc[measures.index(measure)+1] = category.values()


    # df_result=pd.DataFrame(counts).T
    output_file = REPORTPATH +"/summary_report.xlsx"
    with pd.ExcelWriter(output_file) as writer:
        df_category.to_excel(writer, sheet_name="Category Level", index=1)
        df_results.to_excel(writer, sheet_name='File Level Accuracy', index=1)
        print(f"Summary report Generated here : {os.path.abspath(output_file)}")
        shutil.rmtree("temp")

# analyze_summary_report()