import logging

import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


class ConditionalChecks:


    def __init__(self, datafile, lookupfile):
        self.df_datafile=pd.read_excel(datafile)
        self.df_conditional_lookup = pd.read_excel(lookupfile)

    def columns_to_lowercase(self):
        self.df_datafile.columns.str.lower()  # Convert to lowercase
        self.df_conditional_lookup.columns.str.lower()
        self.df_datafile.columns = self.df_datafile.columns.str.replace(r'[^a-z0-9]', '_', regex=True)
        self.df_conditional_lookup.columns = self.df_conditional_lookup.columns.str.replace(r'[^a-z0-9]', '_', regex=True)

    def highlight_and_add_comments(self, ws, row, col, message, color_fill):
        cell = ws.cell(row=row, column=col)
        cell.fill = color_fill
        cell.comment = openpyxl.comments.Comment(message, "Validation Script")

    def verify_original_name_data(self, report):
        # verifying in the report listed columns expected values are matching or not
        wb = load_workbook(report)
        ws = wb.active
        self.df_datafile = self.df_datafile.applymap(lambda x: x.lower() if isinstance(x, str) else x)
        self.df_conditional_lookup = self.df_conditional_lookup.applymap(lambda x: x.lower() if isinstance(x, str) else x)
        # verify is there a null value
        # Add a comments column for the reasons (in the last column)
        comments_column_index = len(self.df_datafile.columns) + 1
        # Define the color fill for highlighting invalid cells and nulls
        invalid_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        null_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")

        # Validate columns from data file against lookup file
        for i, column in enumerate(self.df_datafile.columns, 1):
            # Modify here if we need to check only for mandatory columns
            if column in self.df_conditional_lookup.columns:
                lookup_values = self.df_conditional_lookup[column].dropna().values  # Get non-null values from lookup column
                for j, value in enumerate(self.df_datafile[column], 2):  # Data rows start from 2 (1 is header)
                    updated_value=''
                    if pd.isna(value):  # Check for null values
                        message =f'NULL value found in {column}'
                        self.highlight_and_add_comments(ws, j, i, "Null value found", null_fill)
                        cell_val = ws.cell(row=j, column=comments_column_index, value=message)
                        existing_value = cell_val.value
                        new_value = column
                        if existing_value:
                            updated_value = f"{existing_value} {new_value}"  # You can use a different separator like a comma
                        else:
                            updated_value = new_value
                        cell_val = updated_value
                    elif value not in lookup_values:  # Check if value is not in the lookup values
                        message =f'Data not matching with lookup in {column}'
                        self.highlight_and_add_comments(ws, j, i, "Value not found in lookup", invalid_fill)
                        cell_val = ws.cell(row=j, column=comments_column_index, value=message)
                        existing_value = cell_val.value
                        new_value = column
                        if existing_value:
                            updated_value = f"{existing_value} {new_value}"  # You can use a different separator like a comma
                        else:
                            updated_value = new_value
                        cell_val = updated_value

            else:
                logging.info(f"Column '{column}' is not checked conditional field verification.")
        wb.save(report)

    def verify_for_non_negative(self, report):
        columns = ["Quantity", "Total_Price", "totalprice"]
        wb = load_workbook(report)
        ws = wb.active
        fill_for_invalid = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
        for column in columns:
            if column in self.df_datafile.columns:
                col_index = self.df_datafile.columns.get_loc(column)+1

                for row_id, value in enumerate(self.df_datafile[column], start=2):
                    if value<0:
                        self.highlight_and_add_comments(ws, row_id, col_index, "negative_value", fill_for_invalid)
        wb.save(report)
