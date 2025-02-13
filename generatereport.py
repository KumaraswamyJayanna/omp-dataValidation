from datetime import datetime

import pandas as pd
import openpyxl
import os
from openpyxl.styles import PatternFill
from config import REPORTPATH


class ExcelReport:
    def __init__(self, file1_path, file2_path):
        self.file1_path = file1_path
        self.file2_path = file2_path
        self.df1 = pd.read_excel(self.file2_path)  # Pipeline output
        self.df2 = pd.read_excel(self.file1_path)  # GT output
        self.df1 = self.df1.astype(str)
        self.df2 = self.df2.astype(str)
        self.varvalue = os.path.splitext(os.path.basename(self.file1_path))[0]
        self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.report_path = REPORTPATH +f"/highlighted_report_{self.varvalue + self.timestamp}.xlsx"
        self.sheets_name = ['Pipeline', 'MissingRows', "ExtraRowsinGT"]

    def create_report_sheet(self):
        workbook = openpyxl.Workbook()
        if 'Sheet' in workbook.sheetnames:
            std = workbook['Sheet']
            workbook.remove(std)

        for sheetname in self.sheets_name:
            workbook.create_sheet(title=sheetname)
        workbook.save(self.report_path)

    def append_data_to_report_highlight(self, sheetname, data, columns_to_highlight):
        workbook = openpyxl.load_workbook(self.report_path)
        if sheetname in workbook.sheetnames:
            sheet = workbook[sheetname]
            sheet.append(data)
        else:
            print(f"Sheet '{sheetname}' does not exist in the workbook.")
            return

        last_row = sheet.max_row
        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        if columns_to_highlight:
            for column in columns_to_highlight:
                cell = sheet.cell(row=last_row, column=column)
                cell.fill = fill
        workbook.save(self.report_path)

    def highlight_complete_row(self, sheetname):
        workbook = openpyxl.load_workbook(self.report_path)
        sheet = workbook[sheetname]
        last_row = sheet.max_row
        fill_for_Gt_not_found = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')
        for cell in sheet[last_row]:
            cell.fill = fill_for_Gt_not_found
        workbook.save(self.report_path)

    def generate_report(self):
        self.create_report_sheet()
        columns_data = self.df1.columns.tolist()
        for sheet in self.sheets_name:
            self.append_data_to_report_highlight(sheetname=sheet, data=columns_data, columns_to_highlight=None)

        for i in range(len(self.df1)):
            pipeline_row_data = self.df1.iloc[i].tolist()
            key = pipeline_row_data[0]
            gt_row_dataset_for_key = self.df2[self.df2['Pseudo_column'] == str(key)].values.tolist()

            if not gt_row_dataset_for_key:
                print(f"DATA NOT FOUND IN GT: No data found in GT for {key}")
                self.highlight_complete_row(sheetname="Pipeline")
                self.append_data_to_report_highlight(sheetname="MissingRows",
                                                     data=pipeline_row_data,
                                                     columns_to_highlight=None)
                continue

            gt_row_dataset_for_key = self.df2[self.df2['Pseudo_column'] == key]
            gt_data_index_values = gt_row_dataset_for_key.index.tolist()
            # print(f"{len(gt_row_dataset_for_key)} number of rows retrieved for {key} has Groundtruth index {gt_data_index_values}")
            gt_row_dataset_for_key = gt_row_dataset_for_key.values.tolist()
            data_differences = []

            if len(gt_row_dataset_for_key) == 1:
                if pipeline_row_data == gt_row_dataset_for_key[0]:
                    # print("Exact match found in comparison in first value")
                    self.df2.drop(gt_data_index_values, inplace=True)
                else:
                    differences = [index + 1 for index, (a, b) in enumerate(zip(pipeline_row_data, gt_row_dataset_for_key[0])) if a != b]
                    self.append_data_to_report_highlight(sheetname="Pipeline", data=pipeline_row_data, columns_to_highlight=differences)
                    self.df2.drop(gt_data_index_values, inplace=True)
            else:
                for data in gt_row_dataset_for_key:
                    print("Verifying for index match with groundtruth data")
                    differences = [index + 1 for index, (a, b) in enumerate(zip(pipeline_row_data, data)) if a != b]
                    data_differences.append(differences)

                # print(f"Value Differences found for key {data_differences}")
                if not data_differences:
                    print("Exact matches found")
                    self.append_data_to_report_highlight(sheetname="Pipeline", data=pipeline_row_data, columns_to_highlight=None)
                    self.df2.drop(gt_data_index_values, inplace=True)
                else:
                    original_index = dict(zip(gt_data_index_values, data_differences))
                    find_minimum_difference = min(data_differences, key=len)
                    for index, difference in original_index.items():
                        if difference == find_minimum_difference:
                            val = index
                    min_length_index = data_differences.index(find_minimum_difference)
                    self.append_data_to_report_highlight(sheetname="Pipeline", data=pipeline_row_data, columns_to_highlight=data_differences[min_length_index])
                    self.df2.drop(val, inplace=True)

        gt_data_set = self.df2.values.tolist()
        print(f"Left GT Data is of length :: {len(gt_data_set)}")
        for gt_data in gt_data_set:
            self.append_data_to_report_highlight(sheetname="ExtraRowsinGT", data=gt_data, columns_to_highlight=None)

        print(f"Report Generated Here: {os.path.relpath(self.report_path)}")
        print("Execution Done")
